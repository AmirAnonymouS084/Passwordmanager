import openpyxl
from colorama import Fore , init
import random

def add_password(website , url, user, password):
    sh.cell(row=r+1, column=1).value = website
    sh.cell(row=r+1, column=2).value = url
    sh.cell(row=r+1, column=3).value = user
    sh.cell(row=r+1, column=4).value = password
    print(Fore.GREEN + "Your password was saved successfuly")
def build_password(length):
    characters=[1,2,3,4,5,6,7,8,9,0,"!","@","#","$","%",'q', 'w', 'e',
        'r', 't', 'y', 'u', 'i', 'o', 'p', 'l', 'k', 'j', 'h', 'g', 'f', 
        'd', 's', 'a', 'z', 'x', 'c', 'v', 'b', 'n', 'm', 'M', 'N', 'B', 
        'V', 'C', 'X', 'Z', 'A', 'S', 'D', 'F', 'G', 'H', 'J', 'K', 'L', 
        'P', 'O', 'I', 'U', 'Y', 'T', 'R', 'E', 'W', 'Q', '&', '*', '_']
    password = ''
    if length == 0:
        length = 12
    for i in range(length):
        password += str(random.choice(characters))
    print(Fore.GREEN + password)
    return password

def show_passwords(service):
    for row in sh.iter_rows(min_row=2,max_col=1) :
        for cell in row:
            if cell.value ==  service:
                web = sh.cell(row=cell.row, column=1).value
                usr = sh.cell(row=cell.row, column=3).value
                psswd = sh.cell(row=cell.row, column=4).value
                print(Fore.RED + f"Web : {web}", Fore.CYAN + f"Username : {usr}", Fore.GREEN + f"Passwrord : {psswd}", sep= "\t")

init()
book = openpyxl.load_workbook("Passwords.xlsx")
sh = book.get_sheet_by_name("Passwords")
c = sh.max_column
r = sh.max_row
while True:
    try:
        action = int(input(Fore.CYAN + "[1] Add password [2] Get password : "))
    except ValueError:
        print(Fore.RED + "Select [1] or [2]")
        continue
    if action == 1:
        website = input(Fore.BLUE + "Service name : ")
        url = input(Fore.BLUE + "url : ")
        username = input(Fore.BLUE + "Username : ")
        try :
            passlength = int(input(Fore.BLUE + "Length of your password : "))
        except ValueError:
            print(Fore.RED + "Input a number ! ")
            continue
        passwd = build_password(passlength)
        add_password(website , url, username, passwd)
    elif action == 2:
        service = input(Fore.BLUE + "Service name : ")
        show_passwords(service)
    else:
        print(Fore.RED + "Select [1] or [2]")
    book.save("Passwords.xlsx")

